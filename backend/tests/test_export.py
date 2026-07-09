import csv
import io

from app import export
from app.models import Lead, OutreachStatus


def _lead():
    return Lead(no=5, company_en="LED Factory", country="Chile", city="Santiago",
                email="info@x.cl", phone="+56 9 1", website="x.cl", instagram="ledx",
                stage="negotiating", tags="hot,distributor",
                outreach=[OutreachStatus(channel="whatsapp", status="messaged", message_sent_date="2026-07-08")])


def test_build_csv_has_header_and_row():
    data = export.build_csv([_lead()])
    text = data.decode("utf-8-sig")
    rows = list(csv.reader(io.StringIO(text)))
    assert rows[0][:3] == ["no", "company_en", "country"]
    assert "LED Factory" in rows[1]
    assert "negotiating" in rows[1]
    # whatsapp status column reflects messaged
    hdr = rows[0]
    assert rows[1][hdr.index("whatsapp_status")] == "messaged"


def test_build_xlsx_opens():
    import openpyxl
    data = export.build_xlsx([_lead()])
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb.active
    assert ws.cell(row=1, column=1).value == "no"
    assert ws.cell(row=2, column=2).value == "LED Factory"


def test_export_empty_still_has_header():
    rows = list(csv.reader(io.StringIO(export.build_csv([]).decode("utf-8-sig"))))
    assert rows[0][0] == "no"
    assert len(rows) == 1
